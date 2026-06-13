"""Regenerate sample_menu.xlsx (run from the repo root):

    python scripts/make_sample_menu.py

Creates the template admins upload to the bot: required headers, example
rows, and a few images embedded next to their rows to show how photo
import works.
"""
import io
import os
import sys

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageDraw

HEADERS = ["Name", "Category", "Subcategory", "Price", "Quantity",
           "", "Photo (optional)"]

ROWS = [
    ("Espresso",          "Drinks", "Coffee",      3.00, 50),
    ("Cappuccino",        "Drinks", "Coffee",      4.50, 40),
    ("Latte",             "Drinks", "Coffee",      4.50, 40),
    ("Green Tea",         "Drinks", "Tea",         3.00, 30),
    ("Black Tea",         "Drinks", "Tea",         3.00, 30),
    ("Lemonade",          "Drinks", "Beverages",   3.50, 25),
    ("Iced Tea",          "Drinks", "Beverages",   3.50, 25),
    ("Croissant",         "Food",   "Snacks",      2.50, 20),
    ("Granola Snack Bar", "Food",   "Snacks",      2.00, 30),
    ("Cheesecake",        "Food",   "Desserts",    4.00, 12),
    ("Chicken Rice Bowl", "Food",   "Meals/Bowls", 7.50, 10),
]

# row -> (background, accent) for the placeholder drawings
SAMPLE_IMAGES = {2: ("#6F4E37", "#F5EFE6"),   # Espresso
                 3: ("#C49A6C", "#3E2A1F"),   # Cappuccino
                 9: ("#E8C97D", "#8a5a2b")}   # Croissant


def placeholder(bg: str, fg: str, label: str) -> bytes:
    img = Image.new("RGB", (220, 160), bg)
    d = ImageDraw.Draw(img)
    # a simple cup: body + handle + steam — friendly, no fonts needed
    d.rounded_rectangle([60, 70, 140, 130], 12, fill=fg)
    d.arc([130, 80, 165, 115], -90, 90, fill=fg, width=8)
    for x in (80, 100, 120):
        d.arc([x - 6, 35, x + 6, 60], 180, 360, fill=fg, width=4)
    d.text((8, 140), label, fill=fg)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def main(out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Menu"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="6F4E37")
    body_font = Font(name="Arial")

    ws.append(HEADERS)
    for cell in ws[1]:
        if cell.value:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    for r in ROWS:
        ws.append(list(r))
    for row in ws.iter_rows(min_row=2, max_col=5):
        for cell in row:
            cell.font = body_font
    for row in range(2, len(ROWS) + 2):
        ws.cell(row=row, column=4).number_format = "0.00"
        ws.row_dimensions[row].height = 70   # room for pasted photos

    widths = {"A": 22, "B": 12, "C": 14, "D": 9, "E": 10, "F": 3, "G": 34}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    note = ws.cell(row=len(ROWS) + 3, column=1)
    note.value = ("How to use: keep the headers, one item per row. Category must be "
                  "Drinks or Food; Price is optional. Paste each item's photo into "
                  "column G on the SAME row — the bot picks photos up automatically.")
    note.font = Font(name="Arial", italic=True, color="6F4E37")
    ws.merge_cells(start_row=len(ROWS) + 3, start_column=1,
                   end_row=len(ROWS) + 3, end_column=7)
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[len(ROWS) + 3].height = 45

    for row_no, (bg, fg) in SAMPLE_IMAGES.items():
        name = ROWS[row_no - 2][0]
        xl_img = XlImage(io.BytesIO(placeholder(bg, fg, name)))
        ws.add_image(xl_img, f"G{row_no}")

    wb.save(out_path)
    print(f"wrote {out_path}")


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        os.path.dirname(__file__), "..", "sample_menu.xlsx")
    main(os.path.abspath(out))
