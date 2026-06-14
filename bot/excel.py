"""Parse a menu .xlsx: item rows + images embedded in the sheet.

Expected header row (case-insensitive, any order):
    Name | Category | Subcategory | Price | Quantity
Price is optional. Images pasted into the sheet are matched to items by the
row their top-left corner is anchored to.
"""
import io
from dataclasses import dataclass, field

from openpyxl import load_workbook

REQUIRED = {"name", "category", "subcategory", "quantity"}
KNOWN = REQUIRED | {"price"}
CATEGORY_ALIASES = {"drinks": "Drinks", "drink": "Drinks",
                    "food": "Food", "foods": "Food"}


@dataclass
class ParsedMenu:
    rows: list[dict] = field(default_factory=list)       # validated items
    images: dict[int, bytes] = field(default_factory=dict)  # sheet row -> bytes
    errors: list[str] = field(default_factory=list)
    skipped_images: int = 0


def _norm(v) -> str:
    return str(v).strip() if v is not None else ""


def parse_menu_xlsx(data: bytes) -> ParsedMenu:
    result = ParsedMenu()
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True)
    except Exception as e:                                   # noqa: BLE001
        result.errors.append(f"Could not open the file: {e}")
        return result
    ws = wb.active

    # --- header ------------------------------------------------------------
    header = {}
    for col, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1):
        key = _norm(cell.value).lower()
        if key in KNOWN:
            header[key] = col
    missing = REQUIRED - set(header)
    if missing:
        result.errors.append(
            "Отсутствуют столбцы: " + ", ".join(sorted(missing))
            + ". Ожидаемые заголовки: Name, Category, Subcategory, Price, Quantity."
        )
        return result

    # --- item rows -----------------------------------------------------------
    for row_no, row in enumerate(ws.iter_rows(min_row=2), start=2):
        def cell(key):
            return row[header[key] - 1].value if key in header else None

        name = _norm(cell("name")).replace("|", "/")
        if not name:
            continue                                  # blank line — ignore
        others = [_norm(cell(k)) for k in ("category", "subcategory",
                                           "price", "quantity")]
        if not any(others):
            continue                                  # note row — ignore
        category = _norm(cell("category"))
        if category.lower() in CATEGORY_ALIASES:
            category = CATEGORY_ALIASES[category.lower()]
        category = " ".join(category.split())[:50]
        if not category:
            result.errors.append(
                f"Строка {row_no} («{name}»): категория не заполнена — пропущено."
            )
            continue
        subcategory = " ".join(_norm(cell("subcategory")).split())[:50]
        if not subcategory:
            result.errors.append(
                f"Строка {row_no} («{name}»): подкатегория не заполнена — пропущено."
            )
            continue
        try:
            quantity = max(0, int(float(cell("quantity") or 0)))
        except (TypeError, ValueError):
            result.errors.append(
                f"Строка {row_no} («{name}»): количество не является числом — взято 0."
            )
            quantity = 0
        price = None
        raw_price = cell("price")
        if raw_price not in (None, ""):
            try:
                price = round(float(raw_price), 2)
            except (TypeError, ValueError):
                result.errors.append(
                    f"Строка {row_no} («{name}»): цена не является числом — оставлено пустым."
                )
        result.rows.append({
            "row": row_no, "name": name, "category": category,
            "subcategory": subcategory, "price": price, "quantity": quantity,
        })

    # --- embedded images -----------------------------------------------------
    item_rows = {r["row"] for r in result.rows}
    try:
        for img in getattr(ws, "_images", []):
            try:
                anchor = img.anchor
                row_no = anchor._from.row + 1            # 0-based -> sheet row
                if row_no in item_rows:
                    result.images[row_no] = img._data()
                else:
                    result.skipped_images += 1
            except Exception:                            # noqa: BLE001
                result.skipped_images += 1
    except Exception:                                    # noqa: BLE001
        result.errors.append(
            "Не удалось прочитать встроенные изображения — фото можно прикрепить позже "
            "(отправьте фото с названием товара в подписи)."
        )
    return result
